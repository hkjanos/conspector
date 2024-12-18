import os.path
import pandas as pd
import json
import re
import requests
from bs4 import BeautifulSoup
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Alignment
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.drawing.image import Image
from flask import current_app as app
from app.utils.ticket_export import export_vulnerabilities_to_json  # Import the new utility function
from app.vulnerability.exploit_scanner import check_exploit_for_cve  # Import the exploit scanner function

def extract_cve_id_from_page(url):
    """Extract CVE ID from the given Source URL page using web scraping."""
    if url:
        try:
            # Send an HTTP request to fetch the page content
            response = requests.get(url)
            response.raise_for_status()  # Check if the request was successful

            # Parse the HTML content using BeautifulSoup
            soup = BeautifulSoup(response.text, 'html.parser')

            # Try to find the CVE ID on the page (commonly in <title>, <h1>, or other tags)
            # Different sites may have different structures, so adjust accordingly
            title = soup.find('title')
            if title and 'CVE' in title.text:
                match = re.search(r'CVE-\d{4}-\d+', title.text)
                if match:
                    return match.group(0)  # Return the CVE ID found
            else:
                return "N/A"
        except Exception as e:
            print(f"Error scraping CVE ID from URL {url}: {e}")
            return "N/A"  # Return N/A if there's an error or CVE ID isn't found

    return "N/A"  # Return N/A if URL is empty

def convert_cyclonedx_to_excel(cyclonedx_file_path, output_excel_path):
    """Convert CycloneDX SBOM JSON to Excel, including vulnerabilities and additional fields."""
    try:
        # Load CycloneDX JSON file
        with open(cyclonedx_file_path, 'r') as file:
            sbom_data = json.load(file)

        # Extract components and vulnerabilities from the CycloneDX file
        components = sbom_data.get("components", [])
        vulnerabilities = sbom_data.get("vulnerabilities", [])

        # Prepare data for SBOM (components sheet)
        sbom_rows = []
        for component in components:
            name = component.get("name", "N/A")
            version = component.get("version", "N/A")
            type_ = component.get("type", "N/A")
            license_ = component.get("licenses", [{}])[0].get("license", "N/A")
            description = component.get("description", "N/A")

            # Add a row for the component to the SBOM sheet
            sbom_rows.append([name, version, type_, license_, description])

        # Create a DataFrame for the SBOM sheet
        sbom_df = pd.DataFrame(sbom_rows, columns=["Name", "Version", "Type", "License", "Description"])

        # Prepare data for Vulnerabilities (vulnerabilities sheet)
        vuln_rows = []
        for vuln in vulnerabilities:
            vuln_id = vuln.get("id", "N/A")
            vuln_desc = vuln.get("description", "N/A")
            vuln_severity = "N/A"
            vuln_score = "N/A"
            vuln_source_url = "N/A"
            advisory_urls = []

            # Extract severity and score from ratings
            ratings = vuln.get("ratings", [])
            if ratings:
                vuln_severity = ratings[0].get("severity", "N/A")
                vuln_score = ratings[0].get("score", "N/A")

            # Extract source URL
            vuln_source = vuln.get("source", {})
            if vuln_source:
                vuln_source_url = vuln_source.get("url", "N/A")

            # Extract CVE ID from the source URL by scraping the page
            cve_id = extract_cve_id_from_page(vuln_source_url)

            # Use exploit scanner to check for exploits for this CVE ID
            exploit_info = check_exploit_for_cve(cve_id)  # Check for exploits

            # Extract advisories URLs
            advisories = vuln.get("advisories", [])
            for advisory in advisories:
                advisory_urls.append(advisory.get("url", "N/A"))

            # Affected components are listed under "affects"
            affected_components = vuln.get("affects", [])

            # For each affected component, create a row in the vulnerabilities sheet
            for comp in affected_components:
                affected_component = comp.get("ref", "N/A")
                # For each advisory, create a row with the additional fields
                for advisory_url in advisory_urls:
                    vuln_rows.append(
                        [vuln_id, cve_id, vuln_desc, vuln_severity, vuln_score, affected_component, vuln_source_url,
                         advisory_url, exploit_info])  # Add exploit details to the row

        # Create a DataFrame for the Vulnerabilities sheet with CVE ID between columns A and B
        vuln_df = pd.DataFrame(vuln_rows,
                               columns=["Vulnerability ID", "CVE ID", "Description", "Severity", "CVSS",
                                        "Affected Component", "Source", "Advisory", "Exploit Info"])

        # Write both DataFrames to an Excel file with two separate sheets
        with pd.ExcelWriter(output_excel_path, engine='openpyxl') as writer:
            sbom_df.to_excel(writer, sheet_name='SBOM', index=False)
            vuln_df.to_excel(writer, sheet_name='Vulnerabilities', index=False)

        # Call the utility function to export vulnerabilities to JSON
        ticket_export_json_path = f"{os.path.dirname(output_excel_path)}/ticket_export.json"
        export_vulnerabilities_to_json(vuln_rows, ticket_export_json_path)

        app.logger.debug(f"Excel file saved at: {output_excel_path}")
        app.logger.debug(f"JSON export file saved at: {ticket_export_json_path}")

        # Load the workbook to apply additional formatting
        wb = load_workbook(output_excel_path)

        # Add traffic light coloring to severity column in the Vulnerabilities sheet
        vuln_sheet = wb['Vulnerabilities']
        severity_col = vuln_sheet['D']  # The "Severity" column is the fourth column (now after CVE ID)
        for cell in severity_col[1:]:  # Skip header
            if cell.value == 'high':
                cell.fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")  # Red for high
            elif cell.value == 'medium':
                cell.fill = PatternFill(start_color="FFFF00", end_color="FFFF00",
                                        fill_type="solid")  # Yellow for medium
            elif cell.value == 'low':
                cell.fill = PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid")  # Green for low

        # Convert Source URL and Advisory URLs to hyperlinks in the Vulnerabilities sheet
        for row in range(2, len(vuln_rows) + 2):  # Starting from row 2 (skipping header)
            # Hyperlink the "Source URL"
            source_cell = vuln_sheet.cell(row=row, column=7)
            source_url = source_cell.value
            if source_url != "N/A":
                source_cell.hyperlink = source_url
                source_cell.value = "Source"

            # Hyperlink the "Advisory URL"
            advisory_cell = vuln_sheet.cell(row=row, column=8)
            advisory_url = advisory_cell.value
            if advisory_url != "N/A":
                advisory_cell.hyperlink = advisory_url
                advisory_cell.value = "Advisory"

        # Adjust column widths to fit content, excluding the Description column (which is fixed width)
        for sheet in wb.sheetnames:
            sheet_obj = wb[sheet]
            for col in sheet_obj.columns:
                max_length = 0
                column = col[0].column_letter  # Get the column name
                if column == 'C':  # Skip Description column (C)
                    continue
                for cell in col:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(cell.value)
                    except:
                        pass
                adjusted_width = (max_length + 2)
                sheet_obj.column_dimensions[column].width = adjusted_width

        # Make the Description column (column C) wider and wrap text
        for col in ['C','I']:
            vuln_sheet.column_dimensions[col].width = 80  # Set the width of the Description column to be 80
        #for cell in vuln_sheet[col]:  # Apply text wrapping for all cells in the Description column
            #cell.alignment = Alignment(wrap_text=True, vertical='top')  # Apply top alignment

        # Apply top alignment to all cells
        for sheet_name in wb.sheetnames:
            sheet = wb[sheet_name]
            for row in sheet.iter_rows():
                for cell in row:
                    cell.alignment = Alignment(vertical='top', wrap_text=True)  # Apply top alignment & wrap text

        # Save the workbook with the new formatting
        wb.save(output_excel_path)

    except Exception as e:
        print(f"Error while converting CycloneDX to Excel: {e}")

# Example usage:
# convert_cyclonedx_to_excel('path_to_your_cyclonedx_file.json', 'output_file.xlsx')